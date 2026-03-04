"""Finance reconciliation utility with a Streamlit web UI.

This module lets users upload Excel files, configure comparison options,
and download reconciliation results. The resulting comparison is written
to an Excel file and displayed on screen.
"""

import base64
import json
import os
import tkinter as tk
from dataclasses import dataclass, field
from io import BytesIO
from tkinter import filedialog

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from utils import apply_global_styles, render_header

PROFILES_FILE = "config_profiles.json"

# ── Official Definian brand colors ──────────────────────────────────────── #
_C = {
    "midnight":  "#02072D",
    "blue":      "#0D2C71",
    "green":     "#00AB63",
    "cool_gray": "#D8D7EE",
    "dark_gray": "#3C405B",
    "panel":     "#06103d",
    "text_sec":  "#8a8eb8",
    "border":    "#3C405B",
    "red":       "#e05252",
    "amber":     "#f59e0b",
}


@dataclass
class ComparisonResult:
    """Simple container for comparison output."""

    merged: pd.DataFrame
    summary_text: str
    total_records: int = 0
    matched_records: int = 0
    match_percentage: float = 0.0


def load_profiles() -> dict:
    """Return all saved configuration profiles from disk."""
    if not os.path.exists(PROFILES_FILE):
        return {}
    with open(PROFILES_FILE) as f:
        return json.load(f)


def save_profile(name: str, config: dict) -> None:
    """Persist a named configuration profile to disk."""
    profiles = load_profiles()
    profiles[name] = config
    with open(PROFILES_FILE, "w") as f:
        json.dump(profiles, f, indent=2)


def delete_profile(name: str) -> None:
    """Remove a named configuration profile from disk."""
    profiles = load_profiles()
    profiles.pop(name, None)
    with open(PROFILES_FILE, "w") as f:
        json.dump(profiles, f, indent=2)


def compare_data(
    legacy: pd.DataFrame,
    converted: pd.DataFrame,
    pk_legacy: list[str],
    pk_converted: list[str],
    match_col_legacy: str,
    match_col_converted: str,
    output_file: str | None = None,
    tolerance_type: str | None = None,
    tolerance_value: float | None = None,
    distinct_list: bool = True,
) -> ComparisonResult:
    """Compare two dataframes and return a summary."""

    def _unique_key_name() -> str:
        base = "_comparison_key"
        key_name = base
        counter = 1
        while key_name in legacy.columns or key_name in converted.columns:
            key_name = f"{base}_{counter}"
            counter += 1
        return key_name

    def _build_comparison_key(data: pd.DataFrame, columns: list[str]) -> pd.Series:
        parts = data[columns].fillna("").astype(str)
        return parts.apply(lambda row: " | ".join(value.strip() for value in row), axis=1)

    comparison_key = _unique_key_name()
    legacy = legacy.copy()
    converted = converted.copy()
    legacy_columns = set(legacy.columns)
    converted_columns = set(converted.columns)
    legacy[comparison_key] = _build_comparison_key(legacy, pk_legacy)
    converted[comparison_key] = _build_comparison_key(converted, pk_converted)

    # Convert value columns to numeric now that key columns have been captured as strings
    legacy[match_col_legacy] = pd.to_numeric(legacy[match_col_legacy], errors="coerce")
    converted[match_col_converted] = pd.to_numeric(converted[match_col_converted], errors="coerce")

    if distinct_list:
        legacy = (
            legacy.groupby(comparison_key, dropna=False)
            .agg({match_col_legacy: "sum", **{col: "first" for col in pk_legacy}})
            .reset_index()
        )
        converted = (
            converted.groupby(comparison_key, dropna=False)
            .agg({match_col_converted: "sum", **{col: "first" for col in pk_converted}})
            .reset_index()
        )

    merged_df = pd.merge(
        legacy,
        converted,
        on=comparison_key,
        suffixes=("_legacy", "_converted"),
        how="outer",
    )
    merged_df.rename(columns={comparison_key: "Comparison Key"}, inplace=True)

    legacy_value_column = (
        f"{match_col_legacy}_legacy" if f"{match_col_legacy}_legacy" in merged_df.columns else match_col_legacy
    )
    converted_value_column = (
        f"{match_col_converted}_converted"
        if f"{match_col_converted}_converted" in merged_df.columns
        else match_col_converted
    )

    def _difference(row: pd.Series) -> bool:
        legacy_val = row.get(legacy_value_column)
        conv_val = row.get(converted_value_column)
        if pd.isna(legacy_val) or pd.isna(conv_val):
            return True
        if tolerance_type == "Dollar ($)":
            if tolerance_value is None:
                return legacy_val != conv_val
            return abs(legacy_val - conv_val) > tolerance_value
        if tolerance_type == "Percentage (%)":
            if tolerance_value is None or legacy_val == 0:
                return legacy_val != conv_val
            return abs((legacy_val - conv_val) / legacy_val) > tolerance_value / 100
        return legacy_val != conv_val

    merged_df["Difference"] = merged_df.apply(_difference, axis=1)
    merged_df["Dollar Difference"] = merged_df.apply(
        lambda row: (
            (0 if pd.isna(row.get(legacy_value_column)) else row.get(legacy_value_column))
            - (0 if pd.isna(row.get(converted_value_column)) else row.get(converted_value_column))
        ),
        axis=1,
    )
    merged_df["Percentage Difference"] = merged_df.apply(
        lambda row: (
            float("nan")
            if pd.isna(row.get(legacy_value_column)) or pd.isna(row.get(converted_value_column))
            else (
                abs(row.get(legacy_value_column) - row.get(converted_value_column))
                / abs(row.get(legacy_value_column))
                if row.get(legacy_value_column) != 0
                else float("nan")
            )
        ),
        axis=1,
    )

    def _should_drop_match_column(column_name: str) -> bool:
        if column_name in pk_legacy or column_name in pk_converted:
            return True
        if column_name.endswith("_legacy") and column_name[: -len("_legacy")] in pk_legacy:
            return True
        if column_name.endswith("_converted") and column_name[: -len("_converted")] in pk_converted:
            return True
        return False

    drop_columns = [col for col in merged_df.columns if _should_drop_match_column(col)]
    if drop_columns:
        merged_df.drop(columns=drop_columns, inplace=True)

    rename_columns: dict[str, str] = {}
    for column in merged_df.columns:
        if column in {"Comparison Key", "Difference", "Dollar Difference", "Percentage Difference"}:
            continue
        if column.endswith("_legacy"):
            base = column[: -len("_legacy")]
            rename_columns[column] = f"{base} (First)"
        elif column.endswith("_converted"):
            base = column[: -len("_converted")]
            rename_columns[column] = f"{base} (Second)"
        elif column in legacy_columns and column not in converted_columns:
            rename_columns[column] = f"{column} (First)"
        elif column in converted_columns and column not in legacy_columns:
            rename_columns[column] = f"{column} (Second)"

    if rename_columns:
        merged_df.rename(columns=rename_columns, inplace=True)

    total_records = len(merged_df)
    matched_records = len(merged_df[~merged_df["Difference"]])
    match_percentage = (matched_records / total_records) * 100 if total_records else 0

    summary_lines = [
        f"Total records: {total_records}",
        f"Matched records: {matched_records}",
        f"Match percentage: {match_percentage:.2f}%",
    ]

    if output_file:
        merged_df.to_excel(output_file, index=False)
        _format_output(output_file, merged_df)

    return ComparisonResult(
        merged_df,
        "\n".join(summary_lines),
        total_records=total_records,
        matched_records=matched_records,
        match_percentage=match_percentage,
    )


def _format_output(output_file: str, merged_df: pd.DataFrame) -> None:
    workbook = load_workbook(output_file)
    worksheet = workbook.active
    header_map = {header: idx + 1 for idx, header in enumerate(merged_df.columns)}
    dollar_column = header_map.get("Dollar Difference")
    percent_column = header_map.get("Percentage Difference")

    if dollar_column:
        for row_idx in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_idx, column=dollar_column).number_format = "$#,##0.00"

    if percent_column:
        for row_idx in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_idx, column=percent_column).number_format = "0.00%"

    workbook.save(output_file)


def _section_header(text: str) -> None:
    """Render a branded section label."""
    st.markdown(
        f"""<div style="font-family:'Titillium Web',sans-serif; font-size:10px;
                        font-weight:700; letter-spacing:1.5px; text-transform:uppercase;
                        color:{_C['text_sec']}; border-bottom:1px solid {_C['border']};
                        padding-bottom:6px; margin-bottom:12px; margin-top:4px;">
                {text}
            </div>""",
        unsafe_allow_html=True,
    )


# ─────────────────────────── Streamlit UI ─────────────────────────────── #
def main():
    """Main Streamlit application."""

    st.set_page_config(page_title="GL Reconciliation", layout="wide")
    apply_global_styles()
    render_header("GL Reconciliation")

    # ── Session state ── #
    for key, default in [
        ("first_df", None),
        ("second_df", None),
        ("result", None),
        ("pending_profile", None),
        ("show_save_form", False),
    ]:
        if key not in st.session_state:
            st.session_state[key] = default

    # ── Feedback button — right-aligned, sits at the tab row level ── #
    st.markdown(
        """
        <style>
        .definian-tab-row {
            display: flex;
            align-items: center;
            justify-content: flex-end;
            margin-bottom: -44px;   /* overlap with the tab bar below */
            position: relative;
            z-index: 10;
            pointer-events: none;   /* let clicks fall through to tabs */
        }
        .definian-feedback-btn {
            pointer-events: all;
            display: inline-flex;
            align-items: center;
            gap: 6px;
            padding: 6px 16px;
            background-color: #00AB63;
            border-radius: 5px;
            text-decoration: none !important;
            cursor: pointer;
            transition: background-color 0.15s ease;
            white-space: nowrap;
        }
        .definian-feedback-btn:hover {
            background-color: #00c870;
        }
        .definian-feedback-btn span {
            color: #ffffff !important;
            font-family: 'Titillium Web', sans-serif !important;
            font-size: 11px !important;
            font-weight: 700 !important;
            letter-spacing: 0.8px !important;
            text-transform: uppercase !important;
            line-height: 1.2;
        }
        </style>
        <div class="definian-tab-row">
            <a class="definian-feedback-btn"
               href="https://app.smartsheet.com/b/form/019c9b6cf0b47518a512747269be8f97"
               target="_blank" rel="noopener noreferrer">
                <span>&#8599; Submit Feedback</span>
            </a>
        </div>
        """,
        unsafe_allow_html=True,
    )

    recon_tab, config_tab = st.tabs(["Run Recon", "Manage Configuration"])

    # ══════════════════════════════════════════════════════════════════════ #
    #  TAB 1 — Run Recon
    # ══════════════════════════════════════════════════════════════════════ #
    with recon_tab:
        col1, col2 = st.columns(2, gap="large")

        # ── Upload Files ── #
        with col1:
            _section_header("Upload Files")

            first_file = st.file_uploader(
                "First file",
                type=["xlsx", "xls"],
                key="first_file",
                help="Upload the first Excel file to compare",
            )
            if first_file:
                try:
                    st.session_state.first_df = pd.read_excel(first_file, dtype=str)
                    st.success(f"Loaded: {first_file.name}")
                except Exception as e:
                    st.error(f"Error reading first file: {e}")
                    st.session_state.first_df = None

            second_file = st.file_uploader(
                "Second file",
                type=["xlsx", "xls"],
                key="second_file",
                help="Upload the second Excel file to compare",
            )
            if second_file:
                try:
                    st.session_state.second_df = pd.read_excel(second_file, dtype=str)
                    st.success(f"Loaded: {second_file.name}")
                except Exception as e:
                    st.error(f"Error reading second file: {e}")
                    st.session_state.second_df = None

        # ── Configure ── #
        with col2:
            _section_header("Configure")

            if st.session_state.first_df is not None and st.session_state.second_df is not None:
                first_cols = list(st.session_state.first_df.columns)
                second_cols = list(st.session_state.second_df.columns)

                # Load profile shortcut
                profiles = load_profiles()
                if profiles:
                    lp_col1, lp_col2 = st.columns([4, 1])
                    with lp_col1:
                        profile_to_load = st.selectbox(
                            "Load profile",
                            [""] + list(profiles.keys()),
                            key="profile_selector",
                        )
                    with lp_col2:
                        st.markdown("<div style='margin-top:24px'></div>", unsafe_allow_html=True)
                        if st.button("Apply", key="load_profile_btn") and profile_to_load:
                            st.session_state.pending_profile = profiles[profile_to_load]
                            st.rerun()

                # Apply a pending profile now that columns are known
                if st.session_state.pending_profile:
                    cfg = st.session_state.pending_profile
                    st.session_state["match_keys_first"] = [
                        k for k in cfg.get("match_keys_first", []) if k in first_cols
                    ]
                    st.session_state["match_keys_second"] = [
                        k for k in cfg.get("match_keys_second", []) if k in second_cols
                    ]
                    if cfg.get("compare_col_first") in first_cols:
                        st.session_state["compare_col_first"] = cfg["compare_col_first"]
                    if cfg.get("compare_col_second") in second_cols:
                        st.session_state["compare_col_second"] = cfg["compare_col_second"]
                    tolerance_options = ["None", "Dollar ($)", "Percentage (%)"]
                    saved_tol = cfg.get("tolerance_type", "None")
                    if saved_tol in tolerance_options:
                        st.session_state["tolerance_type_select"] = saved_tol
                    if cfg.get("tolerance_value") is not None:
                        st.session_state["tolerance_value_input"] = float(cfg["tolerance_value"])
                    st.session_state.pending_profile = None

                match_keys_first = st.multiselect(
                    "Match keys — first file",
                    first_cols,
                    key="match_keys_first",
                )
                match_keys_second = st.multiselect(
                    "Match keys — second file",
                    second_cols,
                    key="match_keys_second",
                )

                cmp_col1, cmp_col2 = st.columns(2)
                with cmp_col1:
                    compare_col_first = st.selectbox(
                        "Compare column — first file",
                        first_cols,
                        key="compare_col_first",
                    )
                with cmp_col2:
                    compare_col_second = st.selectbox(
                        "Compare column — second file",
                        second_cols,
                        key="compare_col_second",
                    )

                tol_col1, tol_col2 = st.columns(2)
                with tol_col1:
                    tolerance_type = st.selectbox(
                        "Tolerance type",
                        ["None", "Dollar ($)", "Percentage (%)"],
                        key="tolerance_type_select",
                    )
                with tol_col2:
                    tolerance_value = None
                    if tolerance_type != "None":
                        is_pct = tolerance_type == "Percentage (%)"
                        tolerance_value = st.number_input(
                            "Tolerance value (%)" if is_pct else "Tolerance value ($)",
                            min_value=0.0,
                            max_value=100.0 if is_pct else None,
                            value=0.0,
                            step=0.1 if is_pct else 0.01,
                            format="%.1f" if is_pct else "%.2f",
                            help="Enter a percentage, e.g. 10 means 10%" if is_pct else None,
                            key="tolerance_value_input",
                        )

                st.markdown("<div style='margin-top:8px'></div>", unsafe_allow_html=True)
                run_col, save_col = st.columns(2)
                with run_col:
                    if st.button("Run Comparison", use_container_width=True, key="run_btn"):
                        if not match_keys_first or not match_keys_second:
                            st.error("Please select match key columns for both files.")
                        elif not compare_col_first or not compare_col_second:
                            st.error("Please select compare columns for both files.")
                        else:
                            try:
                                with st.spinner("Running comparison..."):
                                    result = compare_data(
                                        st.session_state.first_df,
                                        st.session_state.second_df,
                                        match_keys_first,
                                        match_keys_second,
                                        compare_col_first,
                                        compare_col_second,
                                        output_file=None,
                                        tolerance_type=None if tolerance_type == "None" else tolerance_type,
                                        tolerance_value=tolerance_value if tolerance_type != "None" else None,
                                        distinct_list=True,
                                    )
                                    st.session_state.result = result
                                    st.success("Comparison complete!")
                            except Exception as e:
                                st.error(f"Error during comparison: {e}")
                                st.session_state.result = None
                with save_col:
                    if st.button("Save Configuration", use_container_width=True, key="open_save_btn"):
                        st.session_state.show_save_form = not st.session_state.show_save_form

                # Inline save form
                if st.session_state.show_save_form:
                    st.markdown(
                        f"""<div style="background:{_C['panel']}; border:1px solid {_C['border']};
                                        border-left:3px solid {_C['green']}; border-radius:5px;
                                        padding:0.75rem 1rem; margin-top:0.5rem;">
                                <span style="font-family:'Titillium Web',sans-serif; font-size:10px;
                                             font-weight:700; letter-spacing:1px; text-transform:uppercase;
                                             color:{_C['text_sec']};">Save Current Configuration</span>
                            </div>""",
                        unsafe_allow_html=True,
                    )
                    sf_col1, sf_col2, sf_col3 = st.columns([5, 1, 1])
                    with sf_col1:
                        save_name = st.text_input(
                            "Profile name",
                            key="inline_save_name",
                            placeholder="e.g. Monthly Vendor Reconciliation",
                            label_visibility="collapsed",
                        )
                    with sf_col2:
                        if st.button("Save", key="inline_save_btn"):
                            if not save_name:
                                st.error("Enter a name.")
                            else:
                                save_profile(
                                    save_name,
                                    {
                                        "match_keys_first": match_keys_first,
                                        "match_keys_second": match_keys_second,
                                        "compare_col_first": compare_col_first,
                                        "compare_col_second": compare_col_second,
                                        "tolerance_type": tolerance_type,
                                        "tolerance_value": tolerance_value,
                                    },
                                )
                                st.success(f"\u2018{save_name}\u2019 saved.")
                                st.session_state.show_save_form = False
                                st.rerun()
                    with sf_col3:
                        if st.button("Cancel", key="inline_save_cancel"):
                            st.session_state.show_save_form = False
                            st.rerun()

            else:
                st.info("Upload both files to configure the comparison.")

        # ── Results ── #
        if st.session_state.result:
            result = st.session_state.result

            st.markdown("<div style='margin-top:1.5rem'></div>", unsafe_allow_html=True)
            _section_header("Results")

            # Metric cards
            unmatched = result.total_records - result.matched_records
            pct = result.match_percentage
            pct_color = _C["green"] if pct >= 95 else (_C["amber"] if pct >= 75 else _C["red"])

            card = (
                f"flex:1; background:{_C['panel']}; border-radius:6px;"
                f" padding:1.25rem 1rem; text-align:center;"
                f" border:1px solid {_C['border']};"
            )
            lbl = (
                f"font-family:'Titillium Web',sans-serif; font-size:9px; font-weight:700;"
                f" letter-spacing:1.5px; text-transform:uppercase;"
                f" color:{_C['text_sec']}; margin-bottom:0.5rem;"
            )
            val = (
                "font-family:'Roboto',sans-serif; font-size:2.4rem;"
                " font-weight:700; line-height:1;"
            )

            st.markdown(
                f"""
                <div style="display:flex; gap:0.75rem; margin-bottom:1.5rem;">
                    <div style="{card} border-top:3px solid {_C['blue']};">
                        <div style="{lbl}">Total Records</div>
                        <div style="{val} color:{_C['cool_gray']};">{result.total_records:,}</div>
                    </div>
                    <div style="{card} border-top:3px solid {_C['green']};">
                        <div style="{lbl}">Matched</div>
                        <div style="{val} color:{_C['green']};">{result.matched_records:,}</div>
                    </div>
                    <div style="{card} border-top:3px solid {_C['red']};">
                        <div style="{lbl}">Unmatched</div>
                        <div style="{val} color:{_C['red']};">{unmatched:,}</div>
                    </div>
                    <div style="{card} border-top:3px solid {pct_color};">
                        <div style="{lbl}">Match Rate</div>
                        <div style="{val} color:{pct_color};">{pct:.1f}%</div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

            # Filter
            _section_header("Filter")
            filter_col1, filter_col2 = st.columns(2)
            with filter_col1:
                show_differences_only = st.checkbox("Show differences only", value=False)
            with filter_col2:
                show_matches_only = st.checkbox("Show matches only", value=False)

            filtered_df = result.merged.copy()
            if show_differences_only and not show_matches_only:
                filtered_df = filtered_df[filtered_df["Difference"]]
            elif show_matches_only and not show_differences_only:
                filtered_df = filtered_df[~filtered_df["Difference"]]

            _section_header(f"Preview — {len(filtered_df):,} rows")
            st.dataframe(filtered_df, use_container_width=True, height=550)

            # Download
            output_buffer = BytesIO()
            filtered_df.to_excel(output_buffer, index=False)
            output_buffer.seek(0)

            temp_filename = "temp_reconciliation_results.xlsx"
            with open(temp_filename, "wb") as f:
                f.write(output_buffer.getvalue())
            _format_output(temp_filename, filtered_df)

            with open(temp_filename, "rb") as f:
                excel_data = f.read()

            if st.button("Download Results", key="download_results_btn"):
                root = tk.Tk()
                root.withdraw()
                root.wm_attributes("-topmost", True)
                save_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx")],
                    initialfile="reconciliation_results.xlsx",
                    title="Save reconciliation results",
                )
                root.destroy()
                if save_path:
                    try:
                        with open(save_path, "wb") as f:
                            f.write(excel_data)
                        st.success(f"Saved to {save_path}")
                    except Exception as exc:
                        st.error(f"Could not save file: {exc}")

    # ══════════════════════════════════════════════════════════════════════ #
    #  TAB 2 — Manage Configuration
    # ══════════════════════════════════════════════════════════════════════ #
    with config_tab:
        cfg_col1, cfg_col2 = st.columns(2, gap="large")

        # ── Save Profile ── #
        with cfg_col1:
            _section_header("Save Current Configuration")

            if (
                st.session_state.first_df is not None
                and st.session_state.second_df is not None
                and "match_keys_first" in st.session_state
            ):
                profile_name = st.text_input(
                    "Profile name",
                    key="profile_name_input",
                    placeholder="e.g. Monthly Vendor Reconciliation",
                )
                if st.button("Save Profile", key="save_profile_btn"):
                    if not profile_name:
                        st.error("Enter a profile name before saving.")
                    else:
                        save_profile(
                            profile_name,
                            {
                                "match_keys_first": st.session_state.get("match_keys_first", []),
                                "match_keys_second": st.session_state.get("match_keys_second", []),
                                "compare_col_first": st.session_state.get("compare_col_first"),
                                "compare_col_second": st.session_state.get("compare_col_second"),
                                "tolerance_type": st.session_state.get("tolerance_type_select", "None"),
                                "tolerance_value": st.session_state.get("tolerance_value_input"),
                            },
                        )
                        st.success(f"Profile \u2018{profile_name}\u2019 saved.")
                        st.rerun()
            else:
                st.info("Upload files and configure a reconciliation in the Run Recon tab first.")

        # ── Saved Profiles ── #
        with cfg_col2:
            _section_header("Saved Profiles")
            profiles = load_profiles()

            if not profiles:
                st.info("No profiles saved yet.")
            else:
                for pname, pcfg in profiles.items():
                    with st.container():
                        st.markdown(
                            f"""<div style="background:{_C['panel']}; border:1px solid {_C['border']};
                                            border-left:3px solid {_C['green']};
                                            border-radius:5px; padding:0.75rem 1rem;
                                            margin-bottom:0.5rem;">
                                    <span style="font-family:'Titillium Web',sans-serif;
                                                 font-size:13px; font-weight:600;
                                                 color:{_C['cool_gray']};">{pname}</span>
                                    <div style="font-family:'Roboto',sans-serif; font-size:11px;
                                                color:{_C['text_sec']}; margin-top:4px; line-height:1.6;">
                                        Match keys (first): {', '.join(pcfg.get('match_keys_first', [])) or '—'}<br>
                                        Match keys (second): {', '.join(pcfg.get('match_keys_second', [])) or '—'}<br>
                                        Compare: {pcfg.get('compare_col_first', '—')} / {pcfg.get('compare_col_second', '—')}<br>
                                        Tolerance: {pcfg.get('tolerance_type', 'None')}
                                        {f" @ {pcfg.get('tolerance_value')}" if pcfg.get('tolerance_value') else ""}
                                    </div>
                                </div>""",
                            unsafe_allow_html=True,
                        )
                        if st.button(f"Delete '{pname}'", key=f"del_{pname}"):
                            delete_profile(pname)
                            st.success(f"Profile \u2018{pname}\u2019 deleted.")
                            st.rerun()


if __name__ == "__main__":
    main()
