"""Core reconciliation logic — framework-agnostic, no Streamlit imports."""

from dataclasses import dataclass, field
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook


@dataclass
class ComparisonResult:
    """Simple container for comparison output."""

    merged: pd.DataFrame
    summary_text: str
    total_records: int = 0
    matched_records: int = 0
    match_percentage: float = 0.0


def compare_data(
    legacy: pd.DataFrame,
    converted: pd.DataFrame,
    pk_legacy: list[str],
    pk_converted: list[str],
    match_col_legacy: str,
    match_col_converted: str,
    output_file: str | None = None,
    tolerance_type: str | None = None,
    tolerance_value: float | None = None,
    distinct_list: bool = True,
) -> ComparisonResult:
    """Compare two dataframes and return a ComparisonResult."""

    def _unique_key_name() -> str:
        base = "_comparison_key"
        key_name = base
        counter = 1
        while key_name in legacy.columns or key_name in converted.columns:
            key_name = f"{base}_{counter}"
            counter += 1
        return key_name

    def _build_comparison_key(data: pd.DataFrame, columns: list[str]) -> pd.Series:
        parts = data[columns].fillna("").astype(str)
        return parts.apply(lambda row: " | ".join(value.strip() for value in row), axis=1)

    comparison_key = _unique_key_name()
    legacy = legacy.copy()
    converted = converted.copy()
    legacy_columns = set(legacy.columns)
    converted_columns = set(converted.columns)
    legacy[comparison_key] = _build_comparison_key(legacy, pk_legacy)
    converted[comparison_key] = _build_comparison_key(converted, pk_converted)

    legacy[match_col_legacy] = pd.to_numeric(legacy[match_col_legacy], errors="coerce")
    converted[match_col_converted] = pd.to_numeric(converted[match_col_converted], errors="coerce")

    if distinct_list:
        legacy = (
            legacy.groupby(comparison_key, dropna=False)
            .agg({match_col_legacy: "sum", **{col: "first" for col in pk_legacy}})
            .reset_index()
        )
        converted = (
            converted.groupby(comparison_key, dropna=False)
            .agg({match_col_converted: "sum", **{col: "first" for col in pk_converted}})
            .reset_index()
        )

    merged_df = pd.merge(
        legacy,
        converted,
        on=comparison_key,
        suffixes=("_legacy", "_converted"),
        how="outer",
    )
    merged_df.rename(columns={comparison_key: "Comparison Key"}, inplace=True)

    legacy_value_column = (
        f"{match_col_legacy}_legacy"
        if f"{match_col_legacy}_legacy" in merged_df.columns
        else match_col_legacy
    )
    converted_value_column = (
        f"{match_col_converted}_converted"
        if f"{match_col_converted}_converted" in merged_df.columns
        else match_col_converted
    )

    def _difference(row: pd.Series) -> bool:
        legacy_val = row.get(legacy_value_column)
        conv_val = row.get(converted_value_column)
        if pd.isna(legacy_val) or pd.isna(conv_val):
            return True
        if tolerance_type == "Dollar ($)":
            if tolerance_value is None:
                return legacy_val != conv_val
            return abs(legacy_val - conv_val) > tolerance_value
        if tolerance_type == "Percentage (%)":
            if tolerance_value is None or legacy_val == 0:
                return legacy_val != conv_val
            return abs((legacy_val - conv_val) / legacy_val) > tolerance_value / 100
        return legacy_val != conv_val

    merged_df["Difference"] = merged_df.apply(_difference, axis=1)
    merged_df["Dollar Difference"] = merged_df.apply(
        lambda row: (
            (0 if pd.isna(row.get(legacy_value_column)) else row.get(legacy_value_column))
            - (0 if pd.isna(row.get(converted_value_column)) else row.get(converted_value_column))
        ),
        axis=1,
    )
    merged_df["Percentage Difference"] = merged_df.apply(
        lambda row: (
            float("nan")
            if pd.isna(row.get(legacy_value_column)) or pd.isna(row.get(converted_value_column))
            else (
                abs(row.get(legacy_value_column) - row.get(converted_value_column))
                / abs(row.get(legacy_value_column))
                if row.get(legacy_value_column) != 0
                else float("nan")
            )
        ),
        axis=1,
    )

    def _should_drop_match_column(column_name: str) -> bool:
        if column_name in pk_legacy or column_name in pk_converted:
            return True
        if column_name.endswith("_legacy") and column_name[: -len("_legacy")] in pk_legacy:
            return True
        if column_name.endswith("_converted") and column_name[: -len("_converted")] in pk_converted:
            return True
        return False

    drop_columns = [col for col in merged_df.columns if _should_drop_match_column(col)]
    if drop_columns:
        merged_df.drop(columns=drop_columns, inplace=True)

    rename_columns: dict[str, str] = {}
    for column in merged_df.columns:
        if column in {"Comparison Key", "Difference", "Dollar Difference", "Percentage Difference"}:
            continue
        if column.endswith("_legacy"):
            base = column[: -len("_legacy")]
            rename_columns[column] = f"{base} (First)"
        elif column.endswith("_converted"):
            base = column[: -len("_converted")]
            rename_columns[column] = f"{base} (Second)"
        elif column in legacy_columns and column not in converted_columns:
            rename_columns[column] = f"{column} (First)"
        elif column in converted_columns and column not in legacy_columns:
            rename_columns[column] = f"{column} (Second)"

    if rename_columns:
        merged_df.rename(columns=rename_columns, inplace=True)

    total_records = len(merged_df)
    matched_records = len(merged_df[~merged_df["Difference"]])
    match_percentage = (matched_records / total_records) * 100 if total_records else 0

    summary_lines = [
        f"Total records: {total_records}",
        f"Matched records: {matched_records}",
        f"Match percentage: {match_percentage:.2f}%",
    ]

    if output_file:
        merged_df.to_excel(output_file, index=False)
        format_output(output_file, merged_df)

    return ComparisonResult(
        merged_df,
        "\n".join(summary_lines),
        total_records=total_records,
        matched_records=matched_records,
        match_percentage=match_percentage,
    )


def format_output(output_file: str, merged_df: pd.DataFrame) -> None:
    """Apply currency/percentage number formats to the Excel output."""
    workbook = load_workbook(output_file)
    worksheet = workbook.active
    header_map = {header: idx + 1 for idx, header in enumerate(merged_df.columns)}
    dollar_column = header_map.get("Dollar Difference")
    percent_column = header_map.get("Percentage Difference")

    if dollar_column:
        for row_idx in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_idx, column=dollar_column).number_format = "$#,##0.00"

    if percent_column:
        for row_idx in range(2, worksheet.max_row + 1):
            worksheet.cell(row=row_idx, column=percent_column).number_format = "0.00%"

    workbook.save(output_file)
